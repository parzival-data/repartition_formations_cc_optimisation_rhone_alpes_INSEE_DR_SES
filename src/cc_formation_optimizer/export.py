"""Exports exploitables d'une solution valide."""

from __future__ import annotations

import csv
import json
import shutil
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean, median
from typing import Any

from cc_formation_optimizer.config import OptimizerConfig
from cc_formation_optimizer.domain import Commune
from cc_formation_optimizer.model_builder import ModelBundle
from cc_formation_optimizer.solution_extractor import CommuneAssignment, ExtractedSolution, OpenSession
from cc_formation_optimizer.validation import ValidationReport


class ExportError(ValueError):
    """Erreur empechant la production d'exports exploitables."""


@dataclass(frozen=True)
class ExportResult:
    """Chemins des fichiers d'export produits."""

    sessions_csv: Path
    assignments_csv: Path
    report_markdown: Path
    statistics_json: Path
    used_config_yaml: Path
    xlsx: Path | None


SESSION_COLUMNS = [
    "id_session",
    "code_pivot",
    "nom_pivot",
    "categorie_pivot",
    "rang_m",
    "type_session",
    "territoire_majoritaire",
    "nombre_communes",
    "nombre_CC",
    "capacite_Q",
    "taux_remplissage",
    "places_restantes",
    "population_min",
    "population_max",
    "population_moyenne",
    "population_mediane",
    "temps_trajet_min",
    "temps_trajet_moyen",
    "temps_trajet_median",
    "temps_trajet_max",
    "nombre_PC",
    "nombre_TPC",
    "nombre_CC_PC",
    "nombre_CC_TPC",
    "mixite_TPC_dans_session_PC",
    "d_jm",
    "cout_eligibilite_pivot",
    "objectif_trajet_session",
    "objectif_eligibilite_session",
    "objectif_mixite_session",
    "alert_level",
    "alert_reasons",
]

ASSIGNMENT_COLUMNS = [
    "code_commune",
    "nom_commune",
    "categorie",
    "territoire_EAR",
    "population",
    "logements",
    "nombre_CC",
    "id_session",
    "code_pivot",
    "nom_pivot",
    "type_session",
    "temps_trajet_minutes",
    "is_pivot",
    "is_same_territory_as_pivot",
    "alert_level",
    "alert_reasons",
]


def export_solution(
    solution: ExtractedSolution,
    validation_report: ValidationReport,
    model_bundle: ModelBundle,
    config: OptimizerConfig,
    config_path: str | Path,
    communes: list[Commune],
    output_dir: str | Path | None = None,
) -> ExportResult:
    """Produit les exports finaux uniquement pour une solution valide."""

    if not validation_report.is_valid:
        raise ExportError("La solution n'a pas passe la validation; aucun export exploitable ne sera produit.")

    root = Path(output_dir) if output_dir is not None else Path(config.exports.get("output_dir", "outputs"))
    solutions_dir = root / "solutions"
    reports_dir = root / "reports"
    solutions_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)

    commune_by_id = {commune.commune_id: commune for commune in communes}
    session_rows = build_session_export_rows(solution, model_bundle, config)
    assignment_rows = build_assignment_export_rows(solution, commune_by_id, config)
    statistics = build_statistics(solution, validation_report, config, session_rows, assignment_rows)

    sessions_csv = solutions_dir / "sessions.csv"
    assignments_csv = solutions_dir / "communes_affectees.csv"
    report_markdown = reports_dir / "rapport_solution.md"
    statistics_json = reports_dir / "statistiques_solution.json"
    used_config_yaml = reports_dir / "config_utilisee.yaml"

    _write_csv(sessions_csv, SESSION_COLUMNS, session_rows)
    _write_csv(assignments_csv, ASSIGNMENT_COLUMNS, assignment_rows)
    _write_json(statistics_json, statistics)
    _write_markdown_report(report_markdown, solution, validation_report, config, config_path, session_rows, statistics)
    shutil.copyfile(config_path, used_config_yaml)
    xlsx_path = _write_optional_xlsx(
        solutions_dir / "solution_formations.xlsx",
        session_rows,
        assignment_rows,
        statistics,
        validation_report,
        used_config_yaml,
    )

    return ExportResult(
        sessions_csv=sessions_csv,
        assignments_csv=assignments_csv,
        report_markdown=report_markdown,
        statistics_json=statistics_json,
        used_config_yaml=used_config_yaml,
        xlsx=xlsx_path,
    )


def build_session_export_rows(
    solution: ExtractedSolution,
    model_bundle: ModelBundle,
    config: OptimizerConfig,
) -> list[dict[str, Any]]:
    """Construit les lignes detaillees de sessions ouvertes."""

    assignments_by_session = _assignments_by_session(solution.assignments)
    rows: list[dict[str, Any]] = []
    for session in solution.sessions:
        assignments = assignments_by_session.get(session.id_session, [])
        populations = [assignment.population for assignment in assignments]
        travel_times = [assignment.temps_trajet_minutes for assignment in assignments]
        pc_assignments = [assignment for assignment in assignments if assignment.categorie == "PC"]
        tpc_assignments = [assignment for assignment in assignments if assignment.categorie == "TPC"]
        q = config.parameters.Q
        eligibility_cost = (
            model_bundle.derived.e_j_TPC[session.code_pivot]
            if session.type_session == "TPC"
            else model_bundle.derived.e_j_PC[session.code_pivot]
        )
        travel_objective = sum(assignment.nombre_CC * assignment.temps_trajet_minutes for assignment in assignments)
        reasons = _session_alert_reasons(session, assignments, eligibility_cost, config)
        rows.append(
            {
                "id_session": session.id_session,
                "code_pivot": session.code_pivot,
                "nom_pivot": session.nom_pivot,
                "categorie_pivot": session.categorie_pivot,
                "rang_m": session.rang_m,
                "type_session": session.type_session,
                "territoire_majoritaire": _majority_territory(assignments),
                "nombre_communes": len(assignments),
                "nombre_CC": session.nombre_CC,
                "capacite_Q": q,
                "taux_remplissage": _round_ratio(session.nombre_CC, q),
                "places_restantes": q - session.nombre_CC,
                "population_min": min(populations) if populations else 0,
                "population_max": max(populations) if populations else 0,
                "population_moyenne": round(mean(populations), 2) if populations else 0,
                "population_mediane": median(populations) if populations else 0,
                "temps_trajet_min": min(travel_times) if travel_times else 0,
                "temps_trajet_moyen": round(mean(travel_times), 2) if travel_times else 0,
                "temps_trajet_median": median(travel_times) if travel_times else 0,
                "temps_trajet_max": max(travel_times) if travel_times else 0,
                "nombre_PC": len(pc_assignments),
                "nombre_TPC": len(tpc_assignments),
                "nombre_CC_PC": sum(assignment.nombre_CC for assignment in pc_assignments),
                "nombre_CC_TPC": sum(assignment.nombre_CC for assignment in tpc_assignments),
                "mixite_TPC_dans_session_PC": session.nombre_CC_TPC_dans_session_PC,
                "d_jm": session.d_jm,
                "cout_eligibilite_pivot": eligibility_cost,
                "objectif_trajet_session": travel_objective,
                "objectif_eligibilite_session": eligibility_cost,
                "objectif_mixite_session": session.d_jm,
                "alert_level": _alert_level(reasons),
                "alert_reasons": "; ".join(reasons),
            }
        )
    return rows


def build_assignment_export_rows(
    solution: ExtractedSolution,
    commune_by_id: dict[str, Commune],
    config: OptimizerConfig,
) -> list[dict[str, Any]]:
    """Construit les lignes detaillees de communes affectees."""

    rows: list[dict[str, Any]] = []
    for assignment in solution.assignments:
        commune = commune_by_id[assignment.code_commune]
        pivot = commune_by_id[assignment.code_pivot]
        reasons = _assignment_alert_reasons(assignment, pivot, config)
        rows.append(
            {
                "code_commune": assignment.code_commune,
                "nom_commune": assignment.nom_commune,
                "categorie": assignment.categorie,
                "territoire_EAR": assignment.territoire_EAR or "",
                "population": assignment.population,
                "logements": assignment.logements if assignment.logements is not None else "",
                "nombre_CC": assignment.nombre_CC,
                "id_session": assignment.id_session,
                "code_pivot": assignment.code_pivot,
                "nom_pivot": assignment.nom_pivot,
                "type_session": assignment.type_session,
                "temps_trajet_minutes": assignment.temps_trajet_minutes,
                "is_pivot": assignment.code_commune == assignment.code_pivot,
                "is_same_territory_as_pivot": _same_territory(commune, pivot),
                "alert_level": _alert_level(reasons),
                "alert_reasons": "; ".join(reasons),
            }
        )
    return rows


def build_statistics(
    solution: ExtractedSolution,
    validation_report: ValidationReport,
    config: OptimizerConfig,
    session_rows: list[dict[str, Any]],
    assignment_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Construit les statistiques JSON de la solution."""

    travel_times = [row["temps_trajet_minutes"] for row in assignment_rows]
    warnings = sorted(
        {reason for row in session_rows + assignment_rows for reason in str(row["alert_reasons"]).split("; ") if reason}
    )
    return {
        "solver_status": solution.status,
        "validation_status": "OK" if validation_report.is_valid else "FAILED",
        "objective_total": solution.objective.objectif_total,
        "obj_trajet": solution.objective.Obj_trajet,
        "obj_eligibilite": solution.objective.Obj_eligibilite,
        "obj_mixite": solution.objective.Obj_mixite,
        "nombre_communes": validation_report.total_assignments,
        "nombre_communes_affectees": len(solution.assignments),
        "nombre_CC": validation_report.total_cc,
        "sessions_ouvertes": len(solution.sessions),
        "B": config.parameters.formation_budgets.B,
        "sessions_PC": sum(1 for session in solution.sessions if session.type_session == "PC"),
        "f": config.parameters.formation_budgets.f,
        "sessions_TPC": sum(1 for session in solution.sessions if session.type_session == "TPC"),
        "k": config.parameters.formation_budgets.k,
        "Q": config.parameters.Q,
        "L": config.parameters.L,
        "T": config.parameters.T,
        "temps_moyen_global": round(mean(travel_times), 2) if travel_times else 0,
        "temps_max_global": max(travel_times) if travel_times else 0,
        "sessions_sous_remplies": sum(1 for row in session_rows if row["nombre_CC"] < config.parameters.L),
        "sessions_saturees": sum(1 for row in session_rows if row["nombre_CC"] == config.parameters.Q),
        "violations": [],
        "warnings": warnings,
    }


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_markdown_report(
    path: Path,
    solution: ExtractedSolution,
    validation_report: ValidationReport,
    config: OptimizerConfig,
    config_path: str | Path,
    session_rows: list[dict[str, Any]],
    statistics: dict[str, Any],
) -> None:
    lines = [
        "# Rapport de solution",
        "",
        f"- Date de generation : {datetime.now().isoformat(timespec='seconds')}",
        f"- Fichier de configuration : `{config_path}`",
        f"- Statut solveur : `{solution.status}`",
        f"- Statut validation : `{statistics['validation_status']}`",
        f"- Nombre de communes : {statistics['nombre_communes']}",
        f"- Nombre de communes affectees : {statistics['nombre_communes_affectees']}",
        f"- Nombre total de CC : {statistics['nombre_CC']}",
        f"- Sessions ouvertes : {statistics['sessions_ouvertes']}",
        f"- Sessions PC : {statistics['sessions_PC']} / f={statistics['f']}",
        f"- Sessions TPC : {statistics['sessions_TPC']} / k={statistics['k']}",
        "",
        "## Parametres",
        "",
        f"- T={config.parameters.T}",
        f"- Q={config.parameters.Q}",
        f"- L={config.parameters.L}",
        f"- B={config.parameters.formation_budgets.B}",
        f"- f={config.parameters.formation_budgets.f}",
        f"- k={config.parameters.formation_budgets.k}",
        "",
        "## Objectif",
        "",
        f"- Objectif total : {solution.objective.objectif_total}",
        f"- Obj_trajet : {solution.objective.Obj_trajet}",
        f"- Obj_eligibilite : {solution.objective.Obj_eligibilite}",
        f"- Obj_mixite : {solution.objective.Obj_mixite}",
        "",
        "## Contraintes validees",
        "",
        "| Controle | Statut |",
        "|---|---|",
    ]
    lines.extend(f"| {constraint} | OK |" for constraint in validation_report.checked_constraints)
    lines.extend(
        [
            "",
            "## Alertes metier",
            "",
            *_markdown_alerts(statistics["warnings"]),
            "",
            "## Top 10 - temps de trajet maximal",
            "",
            *_markdown_session_table(sorted(session_rows, key=lambda row: row["temps_trajet_max"], reverse=True)[:10]),
            "",
            "## Top 10 - sessions les moins remplies",
            "",
            *_markdown_session_table(sorted(session_rows, key=lambda row: row["taux_remplissage"])[:10]),
            "",
            "## Top 10 - sessions les plus mixtes",
            "",
            *_markdown_session_table(
                sorted(session_rows, key=lambda row: row["mixite_TPC_dans_session_PC"], reverse=True)[:10]
            ),
            "",
            "## Donnees et cartographie",
            "",
            "- Les donnees reelles peuvent etre preparees avec `prepare-data` avant resolution.",
            "- Les coordonnees latitude/longitude sont optionnelles. Elles sont integrees aux communes si elles sont presentes dans les donnees propres.",
            "- La carte HTML peut etre produite avec `--map` ou regeneree avec `render-map`; les communes sans coordonnees restent dans les exports mais ne sont pas dessinees.",
            "",
            "## Limites connues",
            "",
            "- L'assouplissement hierarchique est disponible via `solve-relaxed`; il conserve les contraintes dures documentees.",
            "- Les superviseurs, disponibilites et contraintes calendaires ne sont pas optimises par le modele actuel.",
            "- Les exports XLSX sont optionnels et dependent de la disponibilite de la dependance locale.",
            "- Une validation metier finale reste necessaire apres la validation algorithmique.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_optional_xlsx(
    path: Path,
    session_rows: list[dict[str, Any]],
    assignment_rows: list[dict[str, Any]],
    statistics: dict[str, Any],
    validation_report: ValidationReport,
    used_config_yaml: Path,
) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "sessions"
    _write_sheet(default_sheet, SESSION_COLUMNS, session_rows)
    _write_sheet(workbook.create_sheet("communes_affectees"), ASSIGNMENT_COLUMNS, assignment_rows)
    _write_key_value_sheet(workbook.create_sheet("statistiques"), statistics)
    _write_key_value_sheet(
        workbook.create_sheet("validation"),
        {
            "is_valid": validation_report.is_valid,
            "total_sessions": validation_report.total_sessions,
            "total_assignments": validation_report.total_assignments,
            "total_cc": validation_report.total_cc,
            "checked_constraints": ", ".join(validation_report.checked_constraints),
        },
    )
    config_sheet = workbook.create_sheet("configuration")
    for index, line in enumerate(used_config_yaml.read_text(encoding="utf-8").splitlines(), start=1):
        config_sheet.cell(row=index, column=1, value=line)
    workbook.save(path)
    return path


def _write_sheet(sheet: Any, columns: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def _write_key_value_sheet(sheet: Any, payload: dict[str, Any]) -> None:
    sheet.append(["cle", "valeur"])
    for key, value in payload.items():
        sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])


def _assignments_by_session(assignments: tuple[CommuneAssignment, ...]) -> dict[str, list[CommuneAssignment]]:
    by_session: dict[str, list[CommuneAssignment]] = {}
    for assignment in assignments:
        by_session.setdefault(assignment.id_session, []).append(assignment)
    return by_session


def _majority_territory(assignments: list[CommuneAssignment]) -> str:
    territories = [assignment.territoire_EAR for assignment in assignments if assignment.territoire_EAR]
    if not territories:
        return ""
    return Counter(territories).most_common(1)[0][0]


def _session_alert_reasons(
    session: OpenSession,
    assignments: list[CommuneAssignment],
    eligibility_cost: int,
    config: OptimizerConfig,
) -> list[str]:
    thresholds = config.exports["alerts"]
    reasons: list[str] = []
    q = config.parameters.Q
    if session.nombre_CC / q >= thresholds["capacity_close_ratio"]:
        reasons.append("session proche de la capacite maximale")
    if session.nombre_CC / q <= thresholds["low_fill_ratio"]:
        reasons.append("session faiblement remplie")
    if session.temps_trajet_max / config.parameters.T >= thresholds["travel_close_ratio"]:
        reasons.append("temps_trajet_max proche de T")
    territories = {assignment.territoire_EAR for assignment in assignments if assignment.territoire_EAR}
    if len(territories) > 1:
        reasons.append("session multi-territoires")
    if session.type_session == "PC" and session.nombre_CC:
        if session.nombre_CC_TPC_dans_session_PC / session.nombre_CC >= thresholds["high_tpc_mix_ratio"]:
            reasons.append("forte mixite TPC dans session PC")
    if eligibility_cost >= thresholds["high_eligibility_cost"]:
        reasons.append("pivot avec cout d'eligibilite eleve")
    populations = [assignment.population for assignment in assignments if assignment.population > 0]
    if populations and min(populations) > 0 and max(populations) / min(populations) >= thresholds["population_dispersion_ratio"]:
        reasons.append("population tres dispersee")
    return reasons


def _assignment_alert_reasons(
    assignment: CommuneAssignment,
    pivot: Commune,
    config: OptimizerConfig,
) -> list[str]:
    thresholds = config.exports["alerts"]
    reasons: list[str] = []
    if not _same_territory_by_value(assignment.territoire_EAR, pivot.territory_ear):
        reasons.append("commune affectee a un pivot d'un territoire different")
    if assignment.temps_trajet_minutes / config.parameters.T >= thresholds["travel_close_ratio"]:
        reasons.append("temps de trajet proche de T")
    return reasons


def _same_territory(commune: Commune, pivot: Commune) -> bool:
    return _same_territory_by_value(commune.territory_ear, pivot.territory_ear)


def _same_territory_by_value(left: str | None, right: str | None) -> bool:
    if not left or not right:
        return False
    return left == right


def _alert_level(reasons: list[str]) -> str:
    return "WARNING" if reasons else "OK"


def _round_ratio(numerator: int, denominator: int) -> float:
    return round(numerator / denominator, 4) if denominator else 0.0


def _markdown_alerts(warnings: list[str]) -> list[str]:
    if not warnings:
        return ["Aucune alerte metier non bloquante."]
    return [f"- {warning}" for warning in warnings]


def _markdown_session_table(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["Aucune session."]
    lines = [
        "| Session | Pivot | Type | CC | Remplissage | Temps max | Mixite TPC | Alertes |",
        "|---|---|---:|---:|---:|---:|---:|---|",
    ]
    for row in rows:
        lines.append(
            "| {id_session} | {nom_pivot} | {type_session} | {nombre_CC} | {taux_remplissage} | "
            "{temps_trajet_max} | {mixite_TPC_dans_session_PC} | {alert_reasons} |".format(**row)
        )
    return lines
